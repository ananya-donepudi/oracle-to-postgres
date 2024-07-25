#!/usr/bin/env python
# coding: utf-8

# In[19]:


import cx_Oracle
from openpyxl import load_workbook
from sqlalchemy import create_engine, MetaData, Table, Column, VARCHAR
import pandas as pd

# Function to create Oracle connection
def create_oracle_connection(db_params):
    try:
        dsn_tns = cx_Oracle.makedsn(db_params['host'], db_params['port'], service_name=db_params['service_name'])
        connection = cx_Oracle.connect(user=db_params['username'], password=db_params['password'], dsn=dsn_tns)
        print("Oracle Database connection established successfully.")
        return connection
    except cx_Oracle.DatabaseError as e:
        print(f"Error connecting to Oracle Database: {e}")
        return None

# Function to check if a table exists in Oracle
def oracle_table_exists(connection, table_name):
    try:
        cursor = connection.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM user_tables WHERE table_name = '{table_name.upper()}'")
        exists = cursor.fetchone()[0] == 1
        cursor.close()
        return exists
    except cx_Oracle.DatabaseError as e:
        print(f"Error checking if table exists in Oracle: {e}")
        return False

# Function to create a table in Oracle
def create_oracle_table(connection, table_name, columns):
    try:
        cursor = connection.cursor()
        columns_def = ', '.join([f'"{col.upper()}" VARCHAR2(4000)' for col in columns])
        cursor.execute(f"CREATE TABLE {table_name.upper()} ({columns_def})")
        connection.commit()
        cursor.close()
        print(f"Table '{table_name}' created successfully in Oracle.")
    except cx_Oracle.DatabaseError as e:
        print(f"Error creating table in Oracle: {e}")

# Function to upsert data into Oracle table
def upsert_data_to_oracle(connection, table_name, columns, data, primary_key):
    try:
        cursor = connection.cursor()

        # Prepare SQL statements for INSERT and UPDATE
        placeholders = ', '.join([f':{i+1}' for i in range(len(columns))])
        insert_sql = f'INSERT INTO {table_name.upper()} ({", ".join(columns)}) VALUES ({placeholders})'
        update_sql = f'UPDATE {table_name.upper()} SET {", ".join([f"{col} = :{i+1}" for i, col in enumerate(columns)])} WHERE {primary_key.upper()} = :{len(columns)+1}'

        for row in data:
            primary_key_value = row[columns.index(primary_key)]

            # Check if the row exists in the database
            cursor.execute(f'SELECT * FROM {table_name.upper()} WHERE {primary_key.upper()} = :1', (primary_key_value,))
            existing_row = cursor.fetchone()

            if existing_row:
                # Compare each column to check for changes
                should_update = False
                for i, col in enumerate(columns):
                    if str(existing_row[i]) != str(row[i]):
                        should_update = True
                        break

                # If any column has changed, perform an UPDATE
                if should_update:
                    cursor.execute(update_sql, row + [primary_key_value])
                    print(f"Row with {primary_key} = {primary_key_value} updated in Oracle.")
            else:
                # If row doesn't exist, perform an INSERT
                cursor.execute(insert_sql, row)
                print(f"New row with {primary_key} = {primary_key_value} inserted into Oracle.")

        connection.commit()
        cursor.close()
        print("Data upserted successfully in Oracle.")
    except cx_Oracle.DatabaseError as e:
        print(f"Error upserting data into Oracle: {e}")

# Function to create PostgreSQL engine
def create_postgresql_engine(username, password, host, port, database):
    connection_string = f'postgresql://{username}:{password}@{host}:{port}/{database}'
    return create_engine(connection_string)

# Function to create PostgreSQL table
def create_postgresql_table(engine, table_name, columns):
    try:
        metadata = MetaData()
        table = Table(table_name, metadata,
                      *[Column(col.upper(), VARCHAR(4000)) for col in columns])
        metadata.create_all(engine)
        print(f"Table '{table_name}' created successfully in PostgreSQL.")
    except Exception as e:
        print(f"Error creating table in PostgreSQL: {e}")

# Function to upsert data into PostgreSQL table
def upsert_data_to_postgresql(engine, table_name, df):
    try:
        df.to_sql(table_name, engine, if_exists='replace', index=False)
        print("Data upserted successfully in PostgreSQL.")
    except Exception as e:
        print(f"Error upserting data into PostgreSQL: {e}")

# Function to load Excel data into Oracle and PostgreSQL
def load_excel_data_into_databases(oracle_params, postgresql_params, table_name, excel_file_path, primary_key):
    try:
        # Load Excel data
        wb = load_workbook(excel_file_path, data_only=True)  # Load workbook with data_only=True to get cell values
        ws = wb.active
        columns = [str(cell).upper() for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]  # Extract column names
        data = [[str(cell) if cell is not None else '' for cell in row] for row in ws.iter_rows(min_row=2, values_only=True)]  # Extract data rows

        # Create Oracle connection
        oracle_connection = create_oracle_connection(oracle_params)

        if oracle_connection:
            # Check if table exists in Oracle and create if necessary
            if not oracle_table_exists(oracle_connection, table_name):
                create_oracle_table(oracle_connection, table_name, columns)
            
            # Upsert data into Oracle table
            upsert_data_to_oracle(oracle_connection, table_name, columns, data, primary_key)
            
            # Close Oracle connection
            oracle_connection.close()
            print("Oracle Database connection closed.")

        # Create PostgreSQL engine
        postgresql_engine = create_postgresql_engine(**postgresql_params)

        # Create PostgreSQL table if not exists
        create_postgresql_table(postgresql_engine, table_name, columns)
        
        # Upsert data into PostgreSQL table
        upsert_data_to_postgresql(postgresql_engine, table_name, pd.DataFrame(data, columns=columns))

    except Exception as e:
        print(f"An error occurred: {e}")

# Database parameters for Oracle
oracle_params = {
    'username': 'SYSTEM',
    'password': '1234',
    'host': 'desktop-nmaqb8r',
    'port': 1521,
    'service_name': 'XE'
}

# Database parameters for PostgreSQL
postgresql_params = {
    'username': 'postgres',
    'password': '1234',
    'host': 'localhost',
    'port': 5432,
    'database': 'postgres'
}

# Define the table name and the path to the Excel file
table_name = 'weather_details'
excel_file_path = r'C:\Users\wissen\Downloads\weather_data.xlsx'
primary_key = 'CITY'  # Assuming 'CITY' is the primary key

# Call the function to create the table and ingest data
load_excel_data_into_databases(oracle_params, postgresql_params, table_name, excel_file_path, primary_key)


# In[ ]:




